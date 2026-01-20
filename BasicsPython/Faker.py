# from faker import Faker
#
# fake_data = Faker()
# print(fake_data.first_name())
# print(fake_data.last_name())
# print(fake_data.name())
# print(fake_data.email())
# print(fake_data.address())
# print(fake_data.city())
# print(fake_data.zipcode())
#
#
# from openpyxl import Workbook
#
# wb = Workbook()   # Opening
# ws = wb.active    # Active
#
# for i in range(1, 51):
#     for j in range(1, 4):
#         ws.cell(row=i, column=1).value = fake_data.name()
#         ws.cell(row=i, column=2).value = fake_data.email()
#         ws.cell(row=i, column=3).value = fake_data.address()
#
# wb.save('FakeEmail.xlsx')


from faker import Faker
from openpyxl import Workbook

data = Faker()

wb = Workbook()
ws = wb.active

for i in range(1, 51):
    for j in range(1, 6):
        ws.cell(row=i, column=1).value = data.name()
        ws.cell(row=i, column=2).value = data.email()
        ws.cell(row=i, column=3).value = data.address()
        ws.cell(row=i, column=4).value = data.city()
        ws.cell(row=i, column=5).value = data.state()

wb.save("FakeData.xlsx")


